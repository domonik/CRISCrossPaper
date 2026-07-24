import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

SRC = "CrossCelltypeThreePanelCombined.tsv"
OUT = "Tables/precision_at_k_by_dataset.xlsx"

FONT_NAME = "Arial"

df = pd.read_csv(SRC, sep="\t")

precision_cols = [c for c in df.columns if c.startswith("precision_at_")]
# keep numeric-k ordering (precision_at_2 before precision_at_10, etc.)
precision_cols = sorted(precision_cols, key=lambda c: int(c.split("_")[-1]))

recall_cols = [
    c for c in df.columns
    if c.startswith("recall_at_") and c.split("_")[-1].isdigit()
]
recall_cols = sorted(recall_cols, key=lambda c: int(c.split("_")[-1]))

# CRISCross has two feature setups -> treat as two separate "models"
df["model_label"] = df["model"]
mask_criscross = df["model"] == "CRISCross"
df.loc[mask_criscross & (df["features"] == "Sequence & AlphaGenome ATAC-seq"), "model_label"] = "CRISCross (Seq+ATAC)"
df.loc[mask_criscross & (df["features"] == "Sequence only"), "model_label"] = "CRISCross (Seq only)"

# Preferred row order: CRISCross variants first, then the remaining models
# in the order they first appear in the file.
other_models = [m for m in df["model"].unique() if m != "CRISCross"]
row_order = ["CRISCross (Seq+ATAC)", "CRISCross (Seq only)"] + list(other_models)

datasets = list(df["dataset"].unique())

DATASET_ORDER = ["k562_deepcrispr", "k562_full", "ipscs_full"]
datasets = [d for d in DATASET_ORDER if d in datasets] + [d for d in datasets if d not in DATASET_ORDER]

DATASET_LABELS = {
    "ipscs_full": "WTC-11",
    "k562_deepcrispr": "K562 (DeepCrispr)",
    "k562_full": "K562 (CRISPRoffT)",
}
dataset_labels_ordered = [DATASET_LABELS.get(d, d) for d in datasets]

wb = Workbook()
wb.remove(wb.active)


def build_sheet(dataset, ddf, metric_cols, metric_abbrev, metric_name, sheet_name):
    # Average metric@k across all guides for this dataset, per model.
    agg = ddf.groupby("model_label")[metric_cols].mean()
    agg = agg.reindex([m for m in row_order if m in agg.index])
    agg = agg.round(3)

    ws = wb.create_sheet(title=sheet_name[:31])  # Excel sheet-name length limit

    # Header row
    ws.cell(row=1, column=1, value="Model").font = Font(name=FONT_NAME, bold=True)
    for j, col in enumerate(metric_cols, start=2):
        k = col.split("_")[-1]
        c = ws.cell(row=1, column=j, value=f"{metric_abbrev}@{k}")
        c.font = Font(name=FONT_NAME, bold=True)
        c.alignment = Alignment(horizontal="center")

    # Data rows
    for i, model_label in enumerate(agg.index, start=2):
        ws.cell(row=i, column=1, value=model_label).font = Font(name=FONT_NAME)
        for j, col in enumerate(metric_cols, start=2):
            val = agg.loc[model_label, col]
            c = ws.cell(row=i, column=j, value=float(val))
            c.number_format = "0.000"
            c.alignment = Alignment(horizontal="center")
            c.font = Font(name=FONT_NAME)

    # Bold the best (highest) value in each metric@k column
    for j, col in enumerate(metric_cols, start=2):
        col_values = agg[col]
        best_val = col_values.max()
        for i, model_label in enumerate(agg.index, start=2):
            if agg.loc[model_label, col] == best_val:
                cell = ws.cell(row=i, column=j)
                cell.font = Font(name=FONT_NAME, bold=True)

    # Cosmetics: freeze the model-name column, widen it, note the source
    ws.freeze_panes = "B2"
    ws.column_dimensions["A"].width = 24
    for j in range(2, len(metric_cols) + 2):
        ws.column_dimensions[get_column_letter(j)].width = 7

    note_row = len(agg.index) + 3
    n_guides = ddf["guide_id"].nunique()
    ws.cell(
        row=note_row, column=1,
        value=(
            f"Each value = mean {metric_name}@k across {n_guides} unique guides "
            f"in '{dataset}'. Bold = best model at that k. Source: {SRC}"
        ),
    ).font = Font(name=FONT_NAME, italic=True, size=9)


MAX_K = max(int(c.split("_")[-1]) for c in precision_cols)
FLAG_FILL = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
FLAG_FONT = Font(name=FONT_NAME, color="FFFFFFFF", bold=True)


def compute_guide_precision_at_r(ddf):
    # Per guide, R = number of positives for that guide. Look up
    # precision_at_R for each (guide, model); if R exceeds the max recorded
    # k, fall back to precision_at_{MAX_K} and flag the row.
    guides = ddf[["guide_id", "n_positives"]].drop_duplicates()
    rows = []
    for guide_id, r in guides.itertuples(index=False):
        r = int(r)
        flagged = r > MAX_K
        k = min(r, MAX_K)
        col_name = f"precision_at_{k}"
        sub = ddf[ddf["guide_id"] == guide_id]
        for _, row in sub.iterrows():
            rows.append({
                "guide_id": guide_id,
                "n_positives": r,
                "model_label": row["model_label"],
                "value": row[col_name],
                "flagged": flagged,
            })
    return pd.DataFrame(rows)


def build_precision_at_r_sheet(dataset, pr_df, sheet_name):
    # Transposed layout: guides across columns, "Model" / "R (n_positives)"
    # plus one row per model going down.
    guides = (
        pr_df[["guide_id", "n_positives"]]
        .drop_duplicates()
        .sort_values(["n_positives", "guide_id"])
        .reset_index(drop=True)
    )
    models_present = [m for m in row_order if m in pr_df["model_label"].unique()]

    ws = wb.create_sheet(title=sheet_name[:31])

    ws.cell(row=1, column=1, value="Model").font = Font(name=FONT_NAME, bold=True)
    ws.cell(row=2, column=1, value="R (n_positives)").font = Font(name=FONT_NAME, bold=True)
    for j, guide_row in enumerate(guides.itertuples(index=False), start=2):
        guide_id, r = guide_row.guide_id, guide_row.n_positives
        c = ws.cell(row=1, column=j, value=guide_id)
        c.font = Font(name=FONT_NAME, bold=True)
        c.alignment = Alignment(horizontal="center")
        r_cell = ws.cell(row=2, column=j, value=r)
        r_cell.font = Font(name=FONT_NAME)
        r_cell.alignment = Alignment(horizontal="center")

    any_flagged = False
    for i, model_label in enumerate(models_present, start=3):
        ws.cell(row=i, column=1, value=model_label).font = Font(name=FONT_NAME)
        for j, guide_row in enumerate(guides.itertuples(index=False), start=2):
            guide_id = guide_row.guide_id
            match = pr_df[(pr_df["guide_id"] == guide_id) & (pr_df["model_label"] == model_label)]
            if match.empty:
                continue
            val = float(match["value"].iloc[0])
            flagged = bool(match["flagged"].iloc[0])
            c = ws.cell(row=i, column=j, value=round(val, 3))
            c.number_format = "0.000"
            c.alignment = Alignment(horizontal="center")
            if flagged:
                c.font = FLAG_FONT
                c.fill = FLAG_FILL
                any_flagged = True
            else:
                c.font = Font(name=FONT_NAME)

    ws.freeze_panes = "B3"
    ws.column_dimensions["A"].width = 26
    for j in range(2, len(guides) + 2):
        ws.column_dimensions[get_column_letter(j)].width = 16

    note_row = len(models_present) + 4
    note = (
        f"Each value = precision@R per guide in '{dataset}', where R is that "
        f"guide's number of positives. Source: {SRC}"
    )
    if any_flagged:
        note += f" Red = R exceeded the max recorded k ({MAX_K}); value shown is precision@{MAX_K}."
    ws.cell(row=note_row, column=1, value=note).font = Font(name=FONT_NAME, italic=True, size=9)


def build_precision_at_r_rank_sheet(dataset, pr_df, sheet_name):
    # Same transposed layout as build_precision_at_r_sheet, but cell values
    # are the model's rank (1 = best precision@R) among models for that guide.
    # Ties share the same rank (min-rank method).
    guides = (
        pr_df[["guide_id", "n_positives"]]
        .drop_duplicates()
        .sort_values(["n_positives", "guide_id"])
        .reset_index(drop=True)
    )
    models_present = [m for m in row_order if m in pr_df["model_label"].unique()]

    pivot = pr_df.pivot(index="guide_id", columns="model_label", values="value")
    ranks = pivot.rank(axis=1, ascending=False, method="min")

    ws = wb.create_sheet(title=sheet_name[:31])

    ws.cell(row=1, column=1, value="Model").font = Font(name=FONT_NAME, bold=True)
    ws.cell(row=2, column=1, value="R (n_positives)").font = Font(name=FONT_NAME, bold=True)
    for j, guide_row in enumerate(guides.itertuples(index=False), start=2):
        guide_id, r = guide_row.guide_id, guide_row.n_positives
        c = ws.cell(row=1, column=j, value=guide_id)
        c.font = Font(name=FONT_NAME, bold=True)
        c.alignment = Alignment(horizontal="center")
        r_cell = ws.cell(row=2, column=j, value=r)
        r_cell.font = Font(name=FONT_NAME)
        r_cell.alignment = Alignment(horizontal="center")

    for i, model_label in enumerate(models_present, start=3):
        ws.cell(row=i, column=1, value=model_label).font = Font(name=FONT_NAME)
        for j, guide_row in enumerate(guides.itertuples(index=False), start=2):
            guide_id = guide_row.guide_id
            if guide_id not in ranks.index or model_label not in ranks.columns:
                continue
            rank_val = ranks.loc[guide_id, model_label]
            if pd.isna(rank_val):
                continue
            c = ws.cell(row=i, column=j, value=int(rank_val))
            c.alignment = Alignment(horizontal="center")
            c.font = Font(name=FONT_NAME, bold=(rank_val == 1))

    ws.freeze_panes = "B3"
    ws.column_dimensions["A"].width = 26
    for j in range(2, len(guides) + 2):
        ws.column_dimensions[get_column_letter(j)].width = 16

    note_row = len(models_present) + 4
    ws.cell(
        row=note_row, column=1,
        value=(
            f"Rank of each model's precision@R per guide in '{dataset}' "
            "(1 = best; ties share the same rank, min-rank method). "
            f"Source: {SRC}"
        ),
    ).font = Font(name=FONT_NAME, italic=True, size=9)


def latex_escape(s):
    return str(s).replace("_", "\\_")


def write_latex_table(agg, out_file, caption, label):
    datasets_cols = list(agg.columns)
    col_spec = "l" + "c" * len(datasets_cols)

    lines = [
        "\\begin{table}[ht]",
        "\\centering",
        f"\\begin{{tabular}}{{{col_spec}}}",
        "\\hline",
        "Model & " + " & ".join(latex_escape(d) for d in datasets_cols) + " \\\\",
        "\\hline",
    ]
    for model_label in agg.index:
        cells = []
        for dataset in datasets_cols:
            val = agg.loc[model_label, dataset]
            text = f"{val:.3f}"
            if val == agg[dataset].max():
                text = f"\\textbf{{{text}}}"
            cells.append(text)
        lines.append(f"{latex_escape(model_label)} & " + " & ".join(cells) + " \\\\")
    lines += [
        "\\hline",
        "\\end{tabular}",
        f"\\caption{{{caption}}}",
        f"\\label{{{label}}}",
        "\\end{table}",
    ]

    with open(out_file, "w") as f:
        f.write("\n".join(lines) + "\n")
    print(f"Saved {out_file}")


def build_precision_at_r_summary_sheet(pr_by_dataset, sheet_name):
    # rows = models, columns = datasets, value = mean precision@R over guides.
    agg = pr_by_dataset.groupby(["model_label", "dataset"])["value"].mean().unstack("dataset")
    agg = agg.reindex([m for m in row_order if m in agg.index])
    agg = agg[[d for d in dataset_labels_ordered if d in agg.columns]]
    agg = agg.round(3)

    ws = wb.create_sheet(title=sheet_name[:31])

    ws.cell(row=1, column=1, value="Model").font = Font(name=FONT_NAME, bold=True)
    for j, dataset in enumerate(agg.columns, start=2):
        c = ws.cell(row=1, column=j, value=dataset)
        c.font = Font(name=FONT_NAME, bold=True)
        c.alignment = Alignment(horizontal="center")

    for i, model_label in enumerate(agg.index, start=2):
        ws.cell(row=i, column=1, value=model_label).font = Font(name=FONT_NAME)
        for j, dataset in enumerate(agg.columns, start=2):
            val = agg.loc[model_label, dataset]
            c = ws.cell(row=i, column=j, value=float(val))
            c.number_format = "0.000"
            c.alignment = Alignment(horizontal="center")
            c.font = Font(name=FONT_NAME)

    # Bold the best (highest) value in each dataset column
    for j, dataset in enumerate(agg.columns, start=2):
        best_val = agg[dataset].max()
        for i, model_label in enumerate(agg.index, start=2):
            if agg.loc[model_label, dataset] == best_val:
                ws.cell(row=i, column=j).font = Font(name=FONT_NAME, bold=True)

    ws.freeze_panes = "B2"
    ws.column_dimensions["A"].width = 24
    for j in range(2, len(agg.columns) + 2):
        ws.column_dimensions[get_column_letter(j)].width = 18

    note_row = len(agg.index) + 3
    ws.cell(
        row=note_row, column=1,
        value=(
            "Each value = mean precision@R across guides per dataset, where R "
            f"is each guide's number of positives. Bold = best model per dataset. Source: {SRC}"
        ),
    ).font = Font(name=FONT_NAME, italic=True, size=9)

    write_latex_table(
        agg,
        "Tables/patR_summary.tex",
        caption=(
            "Mean precision@R across guides per dataset, where $R$ is each "
            "guide's number of positives. Bold = best model per dataset."
        ),
        label="tab:patr_summary",
    )


all_pr = []
for dataset in datasets:
    label = DATASET_LABELS.get(dataset, dataset)
    ddf = df[df["dataset"] == dataset]
    build_sheet(label, ddf, precision_cols, "P", "precision", f"{label[:28]}_P")
    build_sheet(label, ddf, recall_cols, "R", "recall", f"{label[:28]}_R")

    pr_df = compute_guide_precision_at_r(ddf)
    build_precision_at_r_sheet(label, pr_df, f"{label[:26]}_PatR")
    build_precision_at_r_rank_sheet(label, pr_df, f"{label[:20]}_PatR_Rank")
    pr_df["dataset"] = label
    all_pr.append(pr_df)

build_precision_at_r_summary_sheet(pd.concat(all_pr, ignore_index=True), "PatR_Summary")

wb.save(OUT)
print(f"Saved {OUT} with sheets: {wb.sheetnames}")
