import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

SRC = "CrossCelltypeThreePanelCombined.tsv"
FEATURE_MWU_SRC = "Tables/FeatureMWUTable.tsv"
OUT = "Tables/SupplementaryTables.xlsx"

FONT_NAME = "Arial"

# Internal dataset codes -> display names, matching the convention already
# used in plotCrossCelltypeRecallAt90.py / build_precision_tables.py.
DATASET_LABELS = {
    "ipscs_full": "WTC-11",
    "k562_deepcrispr": "K562 (DeepCRISPR)",
    "k562_full": "K562 (CRISPRoffT)",
}
DATASET_ORDER = ["k562_deepcrispr", "k562_full", "ipscs_full"]
DATASET_RANK = {DATASET_LABELS[d]: i for i, d in enumerate(DATASET_ORDER)}

ROUND_COLS = ["positive_rate", "aucpr", "recall_at_precision90", "precision_at_R"]


def build_guide_id_maps(df):
    # Per dataset code -> {guide_sequence: short_id}. Short IDs (G1, G2, ...)
    # are assigned per dataset ordered by (n_positives, guide_id) ascending,
    # exactly matching the guide_id_maps built in plotPrecisionAtRHeatmaps.py
    # so guide numbers agree between that figure and these tables.
    maps = {}
    for dataset_code, ddf in df.groupby("dataset"):
        guides_sorted = (
            ddf[["guide_id", "n_positives"]]
            .drop_duplicates()
            .sort_values(["n_positives", "guide_id"])
            .reset_index(drop=True)
        )
        maps[dataset_code] = {
            row.guide_id: f"G{i + 1}" for i, row in enumerate(guides_sorted.itertuples(index=False))
        }
    return maps


def build_table_s1(df, guide_id_maps):
    precision_cols = [c for c in df.columns if c.startswith("precision_at_")]
    max_k = max(int(c.split("_")[-1]) for c in precision_cols)

    r = df["n_positives"].astype(int)
    if (r > max_k).any():
        raise ValueError(
            f"n_positives exceeds the max recorded k ({max_k}) for some rows; "
            "extend the precision_at_k range in the source data before rerunning."
        )
    # precision@R per row: R = that row's own number of true positives.
    precision_at_r = df.apply(lambda row: row[f"precision_at_{int(row['n_positives'])}"], axis=1)

    out = df.copy()
    out["precision_at_R"] = precision_at_r
    out["guide"] = out.apply(lambda row: guide_id_maps[row["dataset"]][row["guide_id"]], axis=1)
    out["dataset"] = out["dataset"].map(DATASET_LABELS)

    keep = [
        "model", "dataset", "features", "guide",
        "n_positives", "n_samples", "positive_rate",
        "aucpr", "recall_at_precision90", "precision_at_R",
    ]
    out = out[keep].copy()
    out[ROUND_COLS] = out[ROUND_COLS].round(4)
    out[["n_positives", "n_samples"]] = out[["n_positives", "n_samples"]].astype(int)

    out["_dataset_rank"] = out["dataset"].map(DATASET_RANK)
    out["_guide_rank"] = out["guide"].str.removeprefix("G").astype(int)
    out = out.sort_values(["_dataset_rank", "model", "features", "_guide_rank"])
    out = out.drop(columns=["_dataset_rank", "_guide_rank"]).reset_index(drop=True)
    return out


def build_table_s2(df, guide_id_maps):
    rows = []
    for dataset_code, ddf in df.groupby("dataset"):
        id_map = guide_id_maps[dataset_code]
        guides = ddf[["guide_id", "n_positives", "n_samples"]].drop_duplicates()
        for guide_seq, n_pos, n_samp in guides.itertuples(index=False):
            rows.append({
                "dataset": DATASET_LABELS[dataset_code],
                "guide": id_map[guide_seq],
                "guide_sequence": guide_seq,
                "n_positives": int(n_pos),
                "n_samples": int(n_samp),
            })
    out = pd.DataFrame(rows)
    out["_dataset_rank"] = out["dataset"].map(DATASET_RANK)
    out["_guide_rank"] = out["guide"].str.removeprefix("G").astype(int)
    out = out.sort_values(["_dataset_rank", "_guide_rank"])
    out = out.drop(columns=["_dataset_rank", "_guide_rank"]).reset_index(drop=True)
    return out


def write_sheet(wb, title, df, note=None):
    ws = wb.create_sheet(title=title)

    for j, col in enumerate(df.columns, start=1):
        c = ws.cell(row=1, column=j, value=col)
        c.font = Font(name=FONT_NAME, bold=True)
        c.alignment = Alignment(horizontal="center")

    for i, row in enumerate(df.itertuples(index=False), start=2):
        for j, val in enumerate(row, start=1):
            c = ws.cell(row=i, column=j, value=val)
            c.font = Font(name=FONT_NAME)
            c.alignment = Alignment(horizontal="center")

    ws.freeze_panes = "A2"
    for j, col in enumerate(df.columns, start=1):
        width = max(len(str(col)), df[col].astype(str).str.len().max() if len(df) else 0) + 2
        ws.column_dimensions[get_column_letter(j)].width = min(width, 40)

    if note:
        note_row = len(df) + 3
        ws.cell(row=note_row, column=1, value=note).font = Font(name=FONT_NAME, italic=True, size=9)

    return ws


def main():
    df = pd.read_csv(SRC, sep="\t")
    guide_id_maps = build_guide_id_maps(df)

    s1 = build_table_s1(df, guide_id_maps)
    s2 = build_table_s2(df, guide_id_maps)
    s3 = pd.read_csv(FEATURE_MWU_SRC, sep="\t")

    wb = Workbook()
    wb.remove(wb.active)
    write_sheet(
        wb, "Table S1", s1,
        note=(
            "Per-guide model performance on the cross-cell-type off-target test sets. "
            "precision_at_R = precision at rank R, where R is that guide's number of "
            f"true positives (n_positives). 'guide' short IDs match Table S2. Source: {SRC}"
        ),
    )
    write_sheet(
        wb, "Table S2", s2,
        note=(
            "Mapping between each dataset's short guide IDs (used in Table S1 and in "
            "plotPrecisionAtRHeatmaps.py) and the underlying guide sequence, with each "
            f"guide's number of positives and total candidate off-target sites. Source: {SRC}"
        ),
    )
    write_sheet(
        wb, "Table S3", s3,
        note=(
            "Mann-Whitney U test of AlphaGenome ATAC window-summary features (positives vs. "
            f"negatives), ranked by effect size (rank-biserial correlation). Source: {FEATURE_MWU_SRC}"
        ),
    )

    wb.save(OUT)
    print(f"Saved {OUT} with sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()
